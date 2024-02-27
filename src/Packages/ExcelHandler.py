# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(self.file_path)
        self.sheet_name = self.workbook.sheetnames[0]  # 첫 번째 시트의 이름 가져오기
        self.sheet = self.workbook[self.sheet_name]

    def get_all_keywords(self):
        all_keywords = []
        for row_number, row in enumerate(
            self.sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            keyword = row[0]
            if keyword:
                all_keywords.append((row_number, keyword))
        return all_keywords

    def get_incomplete_keywords(self):
        incomplete_keywords = []
        for row_number, row in enumerate(
            self.sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            keyword = row[0]
            if keyword and not row[-2]:
                incomplete_keywords.append(
                    (row_number, keyword)
                )  # 행 번호와 키워드를 함께 추가합니다.
        return incomplete_keywords

    def bulk_update_cells(self, data):
        for row_key, row_data in data.items():
            row_index = int(row_key)
            for column_name, value in row_data.items():
                column_index = self._get_column_index(column_name)
                self.sheet.cell(row=row_index, column=column_index, value=value)

        self.workbook.save(self.file_path)

    def _get_column_index(self, column_name):
        # 컬럼명을 받아 해당 컬럼의 인덱스를 반환
        for col_idx, col_name in enumerate(
            self.sheet.iter_cols(min_row=1, max_row=1, values_only=True), start=1
        ):
            if col_name[0] == column_name:
                return col_idx

        raise ValueError(f"컬럼 '{column_name}'을(를) 시트에서 찾을 수 없습니다.")


# 예시 사용법
# 테스트용 코드
if __name__ == "__main__":

    excelHandler = ExcelHandler(
        "/Users/seongeun/projects/naver_search_volume/naver_search_volume.xlsx",
    )
    result = excelHandler.get_incomplete_keywords()
    # json_data = {"2": {"1월": 213}, "3": {"1월": 424, "2월": 123}}
    # google_sheet_handler.bulk_update_cells(json_data)
